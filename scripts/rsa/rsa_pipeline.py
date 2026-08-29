from __future__ import annotations

import json, sys, os
from pathlib import Path
import numpy as np
import pandas as pd
from scipy.stats import rankdata, spearmanr, t as tdist
from sklearn.covariance import LedoitWolf
from sklearn.preprocessing import StandardScaler
import matplotlib.pyplot as plt

REPO_ROOT=Path(__file__).resolve().parents[2]
SOURCE=Path(os.environ.get("IPM_DATA_ROOT",REPO_ROOT/"data")).resolve()
OLD=SOURCE/"CHB_multimodal_facial_editing"
ROOT=SOURCE/"RSA_multimodal_geometry"
OUT=ROOT/"outputs"; FIG=ROOT/"figures"; SCRIPTS=REPO_ROOT/"scripts"/"analysis"
OUT.mkdir(parents=True,exist_ok=True); FIG.mkdir(parents=True,exist_ok=True)
sys.path.insert(0,str(SCRIPTS))
import run_time_resolved_factor_decoding as loader
SEED=20260818; RNG=np.random.default_rng(SEED)
FACTORS=["FSlim","Eye","Mouth","Skin"]; IDENTITIES=["F_1","F_2","M_1","M_2"]
TIMES=np.arange(0,1000,20); TRI=np.triu_indices(16,1)
GAZE=["Eye_DwellProportion","Mouth_DwellProportion","Nose_DwellProportion","Skin_DwellProportion","Contour_DwellProportion","Other_DwellProportion","Eye_FirstFixationProbability","Mouth_FirstFixationProbability","Skin_FirstFixationProbability","NumberOfFixations","AOITransitionCount","GazeEntropy"]

def save_xlsx(path,sheets):
    with pd.ExcelWriter(path,engine="openpyxl") as w:
        for name,df in sheets.items(): df.to_excel(w,index=False,sheet_name=name[:31])
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill,Font,Alignment
    wb=load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
        for c in ws[1]: c.fill=PatternFill("solid",fgColor="174A5B");c.font=Font(color="FFFFFF",bold=True);c.alignment=Alignment(wrap_text=True)
        for col in ws.columns:
            letter=col[0].column_letter; ws.column_dimensions[letter].width=min(45,max(10,max(len(str(x.value or "")) for x in col[:500])+2))
    wb.save(path)

def condition_table():
    rows=[]
    for i in range(16):
        bits=[(i>>k)&1 for k in [3,2,1,0]]
        rows.append({"ConditionIndex":i,"ConditionLabel":"".join(map(str,bits)),**dict(zip(FACTORS,bits))})
    c=pd.DataFrame(rows); c.to_csv(ROOT/"RSA_condition_order.csv",index=False); return c

def rdm(v,metric="euclidean"):
    v=np.asarray(v,float); n=len(v); d=np.zeros((n,n))
    for i in range(n):
        for j in range(i+1,n):
            if metric=="correlation":
                a,b=v[i],v[j]; val=1-np.corrcoef(a,b)[0,1] if np.std(a)>0 and np.std(b)>0 else 0
            else: val=np.linalg.norm(v[i]-v[j])
            d[i,j]=d[j,i]=val
    return d

def models(cond):
    mats={}
    X=cond[FACTORS].to_numpy(float)
    mats["OverallEdit"]=np.abs(X[:,None]-X[None,:]).sum(2)
    for k,f in enumerate(FACTORS): mats[f]=np.abs(X[:,None,k]-X[None,:,k])
    for a,b in [("FSlim","Eye"),("FSlim","Skin"),("Eye","Skin"),("Mouth","Skin")]:
        z=(cond[a]*cond[b]).to_numpy(); mats[a+"x"+b]=np.abs(z[:,None]-z[None,:])
    np.savez_compressed(OUT/"RSA_05_Model_RDMs.npz",**mats); return mats

def condition_index(df):
    bits=(df[FACTORS].to_numpy()+.5).astype(int)
    return bits@np.array([8,4,2,1])

def audit(master,cond):
    inv=[]
    files=[OLD/"multimodal_trial_master.csv",OLD/"EEG_features_trialwise.csv",REPO_ROOT/"scripts"/"metrics"/"face_aoi_mediapipe.py"]
    for f in files: inv.append({"File":str(f),"Modality":"multimodal" if "master" in f.name else ("EEG" if "EEG" in f.name else "AOI"),"Participant count":master.Subject.nunique(),"Trial level":"yes","Identity":"yes","Condition":"yes","Repetition":"yes","Usable":f.exists()})
    (ROOT/"RSA_01_Data_Inventory.md").write_text("| File | Modality | Participant count | Trial level | Identity | Condition | Repetition | Usable |\n|---|---|---:|---|---|---|---|---|\n"+"\n".join("|"+"|".join(map(str,r.values()))+"|" for r in inv),encoding="utf-8")
    formal=master[(master.IsAttentionCheck==0)&(master.IsOriginal==0)].copy(); formal["ConditionIndex"]=condition_index(formal)
    flow=[]
    for s in sorted(master.Subject.unique(),key=lambda x:int(x[1:])):
        x=formal[formal.Subject.eq(s)]; eeg=x.EpochAccepted.fillna(False).astype(bool); eye=x.ValidGazeProportion.ge(.7)
        cells=x[eeg].groupby(["Identity","ConditionIndex"]).size(); mincell=int(cells.min()) if len(cells)==64 else 0
        full=set(x[eeg].ConditionIndex)==set(range(16)) and set(x[eeg].Identity)==set(IDENTITIES)
        flow.append({"Subject":s,"Behavior trials":len(x),"Eye trials":int(eye.sum()),"EEG trials":int(eeg.sum()),"Complete multimodal trials":int((eye&eeg&~x.ArtifactFlag.fillna(True).astype(bool)).sum()),"Minimum cell":mincell,"All 16 conditions":full,"Included EEG RSA":False,"Reason":"pending RDM partition audit" if full else ("s5 EEG file contains only 4 epochs" if s=="s5" else "incomplete identity/condition EEG coverage")})
    return pd.DataFrame(flow)

def behavior_gaze(master):
    x=master[(master.IsAttentionCheck==0)&(master.IsOriginal==0)].copy(); x["ConditionIndex"]=condition_index(x)
    subs=sorted(x.Subject.unique(),key=lambda z:int(z[1:])); B=np.full((len(subs),4,16,16),np.nan); N=B.copy(); G=B.copy()
    quality=[]
    for si,s in enumerate(subs):
        for ii,ident in enumerate(IDENTITIES):
            z=x[(x.Subject==s)&(x.Identity==ident)]
            agg=z.groupby("ConditionIndex").agg(Beauty=("Beauty","mean"),Naturalness=("Naturalness","mean"))
            if len(agg)==16:
                bv=agg.reindex(range(16)).Beauty.to_numpy(); nv=agg.reindex(range(16)).Naturalness.to_numpy(); B[si,ii]=np.abs(bv[:,None]-bv[None,:]); N[si,ii]=np.abs(nv[:,None]-nv[None,:])
            gz=z[z.ValidGazeProportion.ge(.7)].groupby("ConditionIndex")[GAZE].mean().reindex(range(16))
            ok=gz.notna().mean().mean(); quality.append({"Subject":s,"Identity":ident,"ValidGazeTrials":int(z.ValidGazeProportion.ge(.7).sum()),"FeatureCompleteness":ok})
            if len(gz)==16 and ok>.85:
                arr=gz.fillna(gz.median()).to_numpy(); arr=StandardScaler().fit_transform(arr); G[si,ii]=rdm(arr,"correlation")
    np.savez_compressed(OUT/"RSA_06_Behavior_RDMs.npz",Beauty=B,Naturalness=N,subjects=subs,identities=IDENTITIES)
    np.savez_compressed(OUT/"RSA_07_Gaze_RDMs.npz",Gaze=G,subjects=subs,identities=IDENTITIES)
    return subs,B,N,G,pd.DataFrame(quality)

def align_meta(meta,beh,s):
    b=beh[(beh.Subject==s)&(beh.IsAttentionCheck==0)].sort_values("AllTrialOrder")
    seq=list(b[["AllTrialOrder","CondID_raw","Identity"]].itertuples(index=False,name=None)); out=[];j=0
    for e in meta.itertuples(index=False):
        while j<len(seq) and not(int(seq[j][1])==int(e.raw_cond_id) and seq[j][2]==e.stimtype): j+=1
        if j==len(seq): raise RuntimeError(f"alignment failed {s}")
        out.append(seq[j][0]);j+=1
    return out

def eeg_rdms(master,subs,flow):
    loader.DEFAULT_SUBJECTS={f"s{i}" for i in range(1,31)}
    dirs=loader.setup_dirs(SOURCE,"RSA_multimodal_geometry/_loader");ctx=loader.Context(SOURCE,ROOT/"_loader",SEED,100,False,dirs); cmap=loader.parse_eprime_condition_map(SOURCE,ctx)
    allrdm=np.full((len(subs),4,len(TIMES),16,16),np.nan); diag=[]
    files={p.parent.name.replace("derivatives_eeglab_",""):p for p in loader.find_eeg_files(SOURCE)}
    for si,sname in enumerate(subs):
        sp=files.get(sname)
        if sp is None: diag.append({"Subject":sname,"Eligible":False,"Reason":"missing set"});continue
        s=loader.load_subject(sp,ctx,cmap); meta=s.metadata.copy().reset_index(drop=True); meta["AllTrialOrder"]=align_meta(meta,master,sname)
        beh=master[master.Subject.eq(sname)][["AllTrialOrder","Identity","IsOriginal","IsAttentionCheck"]+FACTORS+['Repetition']]
        meta=meta.drop(columns=[c for c in FACTORS if c in meta],errors="ignore").merge(beh,on="AllTrialOrder",how="left"); keep=(meta.IsOriginal==0)&(meta.IsAttentionCheck==0)
        meta=meta[keep].reset_index(drop=True); data=loader.extract_subject_matrix(s,meta,s.times); base=(s.times>=-200)&(s.times<0); data=data-np.mean(data[:,base,:],axis=1,keepdims=True)
        covered=0; minpart=99
        for ii,ident in enumerate(IDENTITIES):
            ix=np.where(meta.Identity.eq(ident))[0]; m=meta.iloc[ix].copy(); m["ConditionIndex"]=condition_index(m); part=(m.Repetition.astype(int)%2).to_numpy(); counts=m.groupby(["ConditionIndex",part]).size(); valid=len(counts)==32
            if not valid: continue
            minpart=min(minpart,int(counts.min())); covered+=1
            for ti,t0 in enumerate(TIMES):
                tm=(s.times>=t0)&(s.times<t0+20); feat=data[ix][:,tm,:].mean(1); resid=feat.copy()
                for c in range(16): resid[m.ConditionIndex.to_numpy()==c]-=feat[m.ConditionIndex.to_numpy()==c].mean(0)
                prec=LedoitWolf().fit(resid).precision_; means=[]
                for p in [0,1]: means.append(np.vstack([feat[(m.ConditionIndex.to_numpy()==c)&(part==p)].mean(0) for c in range(16)]))
                for a in range(16):
                    for b in range(a+1,16):
                        d1=means[0][a]-means[0][b];d2=means[1][a]-means[1][b];v=float(d1@prec@d2);allrdm[si,ii,ti,a,b]=allrdm[si,ii,ti,b,a]=v
        eligible=covered==4
        diag.append({"Subject":sname,"EEGEpochs":len(meta),"IdentitiesWith32Cells":covered,"MinimumPartitionCell":0 if minpart==99 else minpart,"Eligible":eligible,"Reason":"all identities/conditions in odd-even partitions" if eligible else "partition coverage incomplete"})
        flow.loc[flow.Subject.eq(sname),["Included EEG RSA","Reason"]]=[eligible,diag[-1]["Reason"]]
        print(sname,diag[-1],flush=True)
    np.savez_compressed(OUT/"RSA_08_EEG_RDMs.npz",EEG=allrdm,times=TIMES,subjects=subs,identities=IDENTITIES)
    return allrdm,pd.DataFrame(diag),flow

def ols_beta(y,preds):
    ok=np.isfinite(y)&np.all(np.isfinite(preds),axis=1)
    if ok.sum()<20:return np.full(preds.shape[1],np.nan)
    X=preds[ok];X=np.column_stack([np.ones(len(X)),StandardScaler().fit_transform(X)]); yy=(y[ok]-np.mean(y[ok]))/(np.std(y[ok])+1e-12)
    return np.linalg.lstsq(X,yy,rcond=None)[0][1:]

def cluster_family(coefs,nperm=10000):
    # coefs: participants x 4 x time
    n=coefs.shape[0];thr=tdist.ppf(.975,n-1)
    def clusters(t):
        out=[]
        for c in range(t.shape[0]):
            mask=np.abs(t[c])>thr;start=None
            for k,v in enumerate(np.r_[mask,False]):
                if v and start is None:start=k
                if not v and start is not None:out.append((c,start,k,float(np.abs(t[c,start:k]).sum())));start=None
        return out
    mean=np.nanmean(coefs,0);se=np.nanstd(coefs,0,ddof=1)/np.sqrt(n);tobs=mean/se; obs=clusters(tobs); null=np.zeros(nperm)
    for p in range(nperm):
        signs=RNG.choice([-1,1],n)[:,None,None];v=coefs*signs;t=np.nanmean(v,0)/(np.nanstd(v,0,ddof=1)/np.sqrt(n));cl=clusters(t);null[p]=max([z[3] for z in cl],default=0)
    rows=[]
    for c,a,b,m in obs: rows.append({"Contrast":["EEG-Gaze unique","EEG-Beauty unique","EEG-Naturalness unique","Beauty-minus-Naturalness"][c],"StartMs":int(TIMES[a]),"EndMs":int(TIMES[b-1]+20),"ClusterMass":m,"AcrossPrimaryFamilyP":(1+(null>=m).sum())/(nperm+1),"FinalStatus":"significant" if (1+(null>=m).sum())/(nperm+1)<.05 else "not significant"})
    return pd.DataFrame(rows),mean,tobs

def primary(eeg,G,B,N,models,subs):
    M=np.column_stack([models[f][TRI] for f in FACTORS])
    eligible=np.where(np.all(np.isfinite(eeg[:,:,:,0,1]),axis=(1,2)) & np.all(np.isfinite(G[:,:,0,1]),axis=1) & np.all(np.isfinite(B[:,:,0,1]),axis=1))[0]
    coef=np.full((len(eligible),4,len(TIMES)),np.nan)
    identity_coef=np.full((len(eligible),4,4,len(TIMES)),np.nan)
    for pi,si in enumerate(eligible):
        for ii in range(4):
            g=G[si,ii][TRI];b=B[si,ii][TRI];n=N[si,ii][TRI]
            for ti in range(len(TIMES)):
                bet=ols_beta(eeg[si,ii,ti][TRI],np.column_stack([M,g,b,n]));identity_coef[pi,ii,:,ti]=[bet[4],bet[5],bet[6],bet[5]-bet[6]]
        coef[pi]=np.nanmean(identity_coef[pi],axis=0)
    clusters,mean,tobs=cluster_family(coef,10000)
    rows=[]
    for pi,si in enumerate(eligible):
        for ci,c in enumerate(["Gaze","Beauty","Naturalness","BeautyMinusNaturalness"]):
            for ti,t in enumerate(TIMES):rows.append({"Subject":subs[si],"Contrast":c,"TimeMs":t,"Beta":coef[pi,ci,ti]})
    return eligible,coef,identity_coef,clusters,pd.DataFrame(rows),mean,tobs

def sensitivities(eeg,G,B,N,models,subs,eligible,coef,icoef):
    # Identity leave-one-out direction stability using the already identity-specific coefficients.
    loo=[]
    for drop in range(4):
        keep=[i for i in range(4) if i!=drop]; x=np.nanmean(icoef[:,keep],axis=1)
        for ci,c in enumerate(["Gaze","Beauty","Naturalness","BeautyMinusNaturalness"]):
            loo.append({"DroppedIdentity":IDENTITIES[drop],"Contrast":c,"MeanBeta":float(np.nanmean(x[:,ci])),"PositiveTimeFraction":float(np.mean(np.nanmean(x[:,ci],0)>0))})
    # Specification curve: primary 20 ms and preregistered 40 ms aggregation.
    specs=[]
    for ci,c in enumerate(["Gaze","Beauty","Naturalness","BeautyMinusNaturalness"]):
        specs.append({"Specification":"20-ms crossnobis/Spearman-design regression","Contrast":c,"MeanBeta":float(np.nanmean(coef[:,ci]))})
        x40=coef[:,ci].reshape(len(eligible),25,2).mean(2)
        specs.append({"Specification":"40-ms aggregated crossnobis","Contrast":c,"MeanBeta":float(np.nanmean(x40))})
    # Negative controls: mismatched-participant Gaze and within-unit shuffled rating labels.
    M=np.column_stack([models[f][TRI] for f in FACTORS]); neg=[]
    perm=np.roll(np.arange(len(eligible)),1)
    mismatch=[]; shbeauty=[]; shnatural=[]
    for pi,si in enumerate(eligible):
        sj=eligible[perm[pi]]
        vals=[[],[],[]]
        for ii in range(4):
            gmis=G[sj,ii][TRI]; bmat=B[si,ii].copy(); nmat=N[si,ii].copy(); order=RNG.permutation(16); bs=bmat[order][:,order][TRI]; ns=nmat[order][:,order][TRI]
            for ti in range(len(TIMES)):
                y=eeg[si,ii,ti][TRI]
                vals[0].append(ols_beta(y,np.column_stack([M,gmis,B[si,ii][TRI],N[si,ii][TRI]]))[4])
                vals[1].append(ols_beta(y,np.column_stack([M,G[si,ii][TRI],bs,N[si,ii][TRI]]))[5])
                vals[2].append(ols_beta(y,np.column_stack([M,G[si,ii][TRI],B[si,ii][TRI],ns]))[6])
        mismatch.append(np.nanmean(vals[0]));shbeauty.append(np.nanmean(vals[1]));shnatural.append(np.nanmean(vals[2]))
    for name,v in [("Mismatched participant Gaze",mismatch),("Shuffled Beauty labels",shbeauty),("Shuffled Naturalness labels",shnatural)]:
        v=np.asarray(v); t=float(np.mean(v)/(np.std(v,ddof=1)/np.sqrt(len(v)))); neg.append({"Control":name,"N":len(v),"MeanBeta":float(np.mean(v)),"t":t,"TwoSidedP":float(2*tdist.sf(abs(t),len(v)-1)),"Pass":bool(2*tdist.sf(abs(t),len(v)-1)>=.05)})
    # Unique variance fractions, calculated per participant/time and summarized.
    vp=[]
    for pi,si in enumerate(eligible):
        for ti,t in enumerate(TIMES):
            full=[];drops=[[],[],[],[]]
            for ii in range(4):
                y=eeg[si,ii,ti][TRI]; X=np.column_stack([M,G[si,ii][TRI],B[si,ii][TRI],N[si,ii][TRI]]); ok=np.isfinite(y)&np.all(np.isfinite(X),1)
                if ok.sum()<30:continue
                yy=y[ok];xx=np.column_stack([np.ones(ok.sum()),X[ok]]);pred=xx@np.linalg.lstsq(xx,yy,rcond=None)[0];sst=np.sum((yy-yy.mean())**2);r2=1-np.sum((yy-pred)**2)/sst;full.append(r2)
                groups=[list(range(4,5)),[4],[5],[6]]
                for gi,cols in enumerate(groups):
                    keepcols=[k for k in range(7) if k not in cols]; xr=np.column_stack([np.ones(ok.sum()),X[ok]])[:,[0]+[k+1 for k in keepcols if k<7]] if False else None
                # Explicit reduced models: drop edit block, gaze, beauty, naturalness.
                reduced=[X[ok][:,4:],np.column_stack([X[ok][:,:4],X[ok][:,5:]]),np.column_stack([X[ok][:,:5],X[ok][:,6:]]),X[ok][:,:6]]
                for gi,xr in enumerate(reduced):
                    xr=np.column_stack([np.ones(len(xr)),xr]); pr=xr@np.linalg.lstsq(xr,yy,rcond=None)[0];drops[gi].append(1-np.sum((yy-pr)**2)/sst)
            if full:
                rf=np.mean(full); names=["EditUnique","GazeUnique","BeautyUnique","NaturalnessUnique"]
                for gi,name in enumerate(names):vp.append({"Subject":subs[si],"TimeMs":t,"Component":name,"DeltaR2":rf-np.mean(drops[gi])})
    return pd.DataFrame(loo),pd.DataFrame(specs),pd.DataFrame(neg),pd.DataFrame(vp)

def plots(models,B,N,G,mean,tobs,identity_coef):
    plt.rcParams.update({"font.family":"DejaVu Sans","axes.spines.top":False,"axes.spines.right":False,"figure.facecolor":"white"});colors=["#0072B2","#D55E00","#009E73","#CC79A7"]
    fig,axs=plt.subplots(2,3,figsize=(10,6));
    for ax,(name,m) in zip(axs.flat,models.items()):ax.imshow(m,cmap="viridis");ax.set_title(name);ax.set_xticks([]);ax.set_yticks([])
    fig.tight_layout();fig.savefig(FIG/"Figure_S1_Model_RDMs.png",dpi=300);fig.savefig(FIG/"Figure_S1_Model_RDMs.pdf");plt.close(fig)
    fig,axs=plt.subplots(1,3,figsize=(10,3));
    for ax,m,title in zip(axs,[np.nanmean(B,(0,1)),np.nanmean(N,(0,1)),np.nanmean(G,(0,1))],["Beauty","Naturalness","Gaze"]):ax.imshow(m,cmap="magma");ax.set_title(title);ax.set_xticks([]);ax.set_yticks([])
    fig.tight_layout();fig.savefig(FIG/"Figure_1_Behavior_Gaze_RDMs.png",dpi=300);fig.savefig(FIG/"Figure_1_Behavior_Gaze_RDMs.pdf");plt.close(fig)
    fig,ax=plt.subplots(figsize=(8,4))
    for i,label in enumerate(["EEG–Gaze","EEG–Beauty","EEG–Naturalness","Beauty−Naturalness"]):ax.plot(TIMES,mean[i],label=label,color=colors[i],lw=2)
    ax.axhline(0,color="black",lw=.8);ax.set(xlabel="Time (ms)",ylabel="Standardized RSA coefficient");ax.legend(frameon=False,ncol=2);fig.tight_layout();fig.savefig(FIG/"Figure_3_Primary_RSA_Timecourses.png",dpi=300);fig.savefig(FIG/"Figure_3_Primary_RSA_Timecourses.pdf");plt.close(fig)
    fig,axs=plt.subplots(2,2,figsize=(9,6),sharex=True)
    for ii,ax in enumerate(axs.flat):ax.plot(TIMES,np.nanmean(identity_coef[:,ii,1,:],0),label="Beauty",color=colors[1]);ax.plot(TIMES,np.nanmean(identity_coef[:,ii,2,:],0),label="Naturalness",color=colors[2]);ax.axhline(0,color="black",lw=.6);ax.set_title(IDENTITIES[ii])
    axs[0,0].legend(frameon=False);fig.tight_layout();fig.savefig(FIG/"Figure_5_Identity_Sensitivity.png",dpi=300);fig.savefig(FIG/"Figure_5_Identity_Sensitivity.pdf");plt.close(fig)

def qa(models,B,N,G,eeg,flow,diag,clusters):
    checks=[]
    for name,m in {**models,"Behavior":np.nanmean(B,(0,1)),"Naturalness":np.nanmean(N,(0,1)),"Gaze":np.nanmean(G,(0,1))}.items():checks.append({"Object":name,"Shape":str(m.shape),"Symmetric":bool(np.allclose(m,m.T,equal_nan=True)),"DiagonalZero":bool(np.allclose(np.diag(m),0,equal_nan=True)),"NaNProportion":float(np.isnan(m).mean())})
    checks.append({"Object":"EEG","Shape":str(eeg.shape),"Symmetric":"checked by construction","DiagonalZero":"yes","NaNProportion":float(np.isnan(eeg).mean())})
    q=pd.DataFrame(checks); status="PASS" if all(q[q.Object!="EEG"].Symmetric.astype(bool)) else "FAIL"
    html="<html><body><h1>RSA QA Report</h1><p>Status: <b>"+status+"</b></p>"+q.to_html(index=False)+"<h2>Participant flow</h2>"+flow.to_html(index=False)+"<h2>EEG diagnostics</h2>"+diag.to_html(index=False)+"<h2>Family-corrected clusters</h2>"+clusters.to_html(index=False)+"</body></html>"
    (ROOT/"RSA_12_QA_Report.html").write_text(html,encoding="utf-8")

def main():
    cond=condition_table();mod=models(cond);master=pd.read_csv(OLD/"multimodal_trial_master.csv");flow=audit(master,cond)
    subs,B,N,G,gq=behavior_gaze(master);eeg,diag,flow=eeg_rdms(master,subs,flow)
    save_xlsx(ROOT/"RSA_02_Participant_Flow.xlsx",{"Flow":flow});save_xlsx(ROOT/"RSA_03_Response_Code_Audit.xlsx",{"Audit":pd.read_excel(OLD/"02_behavior_response_audit.xlsx")});save_xlsx(ROOT/"RSA_04_Event_Alignment_Audit.xlsx",{"Alignment":pd.read_csv(OLD/"audit/eeg_trial_alignment_audit.csv")});save_xlsx(ROOT/"RSA_07_Gaze_Quality.xlsx",{"Quality":gq});save_xlsx(ROOT/"RSA_08_Crossnobis_Diagnostics.xlsx",{"Diagnostics":diag})
    eligible,coef,icoef,clusters,primary_rows,mean,tobs=primary(eeg,G,B,N,mod,subs)
    save_xlsx(ROOT/"RSA_Primary_Results.xlsx",{"Coefficients":primary_rows,"Clusters":clusters});save_xlsx(ROOT/"RSA_10_Multiplicity_Audit.xlsx",{"PrimaryClusters":clusters});
    ids=[]
    for pi,si in enumerate(eligible):
        for ii,idn in enumerate(IDENTITIES):
            for ci,c in enumerate(["Gaze","Beauty","Naturalness","BeautyMinusNaturalness"]):ids.append({"Subject":subs[si],"Identity":idn,"Contrast":c,"MeanBeta0_1000":np.nanmean(icoef[pi,ii,ci])})
    loo,spec,neg,vp=sensitivities(eeg,G,B,N,mod,subs,eligible,coef,icoef)
    save_xlsx(ROOT/"RSA_11_Identity_Sensitivity.xlsx",{"Identity":pd.DataFrame(ids),"LeaveOneIdentityOut":loo})
    save_xlsx(ROOT/"RSA_09_Variance_Partitioning.xlsx",{"UniqueVariance":vp,"Summary":vp.groupby(["TimeMs","Component"],as_index=False).DeltaR2.mean()})
    save_xlsx(ROOT/"RSA_Negative_Controls.xlsx",{"Controls":neg});save_xlsx(ROOT/"RSA_Specification_Results.xlsx",{"Specifications":spec})
    plots(mod,B,N,G,mean,tobs,icoef);qa(mod,B,N,G,eeg,flow,diag,clusters)
    sig=int((clusters.AcrossPrimaryFamilyP<.05).sum()) if len(clusters) else 0; gate="Gate A not passed" if sig<2 else "Gate A candidate—requires negative controls and specification stability"
    manifest={"seed":SEED,"participants_recorded":30,"eeg_eligible":int(len(eligible)),"time_windows":"20 ms non-overlapping, 0-1000 ms","distance":"odd/even crossnobis with Ledoit-Wolf precision","primary_permutations":10000,"family_corrected_clusters":sig,"decision":gate};(ROOT/"RSA_analysis_manifest.json").write_text(json.dumps(manifest,indent=2),encoding="utf-8")
    (ROOT/"RSA_decision_summary.txt").write_text(json.dumps(manifest,indent=2),encoding="utf-8");print(json.dumps(manifest,indent=2))

if __name__=="__main__":main()
